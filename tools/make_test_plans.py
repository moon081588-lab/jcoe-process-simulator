#!/usr/bin/env python3
"""계획서 로더 테스트용 샘플 엑셀/CSV 생성 (형식 변형 4종)."""
import csv, datetime, pathlib, random
import openpyxl
random.seed(7)
OUT = pathlib.Path(__file__).resolve().parents[1] / 'testdata'
OUT.mkdir(exist_ok=True)

def rows(n=40, od_inch=False, l_m=False):
    out = []
    for i in range(n):
        od = random.choice([457, 610, 762, 914, 1016, 1219, 1422])
        t  = random.choice([9.3, 12.7, 15.9, 19.05, 25.4, 31.2])
        L  = random.choice([11800, 12802, 18288, 12000])
        q  = random.choice([5, 12, 25, 40, 70, 120])
        d  = datetime.datetime(2026, 4, 1, 8) + datetime.timedelta(hours=6 * i)
        out.append([f"P{26000+i}", round(od/25.4, 1) if od_inch else od, t,
                    round(L/1000, 3) if l_m else L, q, d, d + datetime.timedelta(days=20)])
    return out

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "JCOE"
ws.append(["오더번호", "외경(mm)", "두께(mm)", "길이(mm)", "수량(본)", "계획일", "납기"])
for r in rows(): ws.append(r)
wb.save(OUT / "plan_standard.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "표지"
ws.append(["2026년 4월 포항공장 조관계획서"])
w2 = wb.create_sheet("JCOE 조관계획")
w2.append(["2026년 4월 조관계획서"]); w2.append(["작성: 생산관리팀"]); w2.append([])
w2.append(["No.", "O.D (inch)", "WT", "L (m)", "PCS", "투입일자"])
for r in rows(od_inch=True, l_m=True): w2.append(r[:6])
wb.save(OUT / "plan_multisheet_inch.xlsx")

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "계획"
ws.append(["제번", "외경", "육후", "길이", "본수"])
for i, r in enumerate(rows(25)):
    row = r[:5]
    if i == 3:  row[1] = None
    if i == 7:  row[4] = 0
    if i == 11: row[2] = 999
    if i == 15: row = [None] * 5
    ws.append(row)
wb.save(OUT / "plan_dirty.xlsx")

with open(OUT / "plan.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f); w.writerow(["ORDER", "OD", "THK", "LENGTH", "QTY", "DATE"])
    for r in rows(15): w.writerow([r[0], r[1], r[2], r[3], r[4], r[5].strftime("%Y-%m-%d")])
print("[ok]", OUT)
