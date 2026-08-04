#!/usr/bin/env python3
"""
★JCOE 공정 생산 표준 시간 분석 엑셀에서 계산 엔진용 룩업 테이블을 추출해
data/tables.json 을 생성합니다.

사용:
    pip install openpyxl
    python3 tools/extract_tables.py "★JCOE 공정 생산 표준 시간 분석 - 20251218 (POSTECH 송부).xlsx"

추출 대상
  tackWeld / insideWeld / outsideWeld  : 두께별 WPS 용접속도 [mm/s] + 패스 수
  utCut                                : 1st-UT 두께별 절단속도
  emSpeed / emFeed                     : Edge Miller 고속 Setting 값 · 메인 피딩기 전진거리
  preBenderPitch                       : Pre Bender 성형 피치
  pressX1                              : Press Bender 외경별 X1 Side Press 횟수
  endFacing / endFacingTC              : End-Facing 저속절삭 시간(안전계수 반영) · 공구교체 시간
  hydroFill / hydroConst               : 수압 외경별 충수시간 · 고정 시간 상수
  expanderDie                          : 확관 1호기 인치·두께별 다이 Step Size [mm]
  packingMarking                       : 포장 마킹 사양수 · 관단수
"""
import json
import pathlib
import re
import sys

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "★JCOE 공정 생산 표준 시간 분석 - 20251218 (POSTECH 송부).xlsx"
OUT = pathlib.Path(__file__).resolve().parents[1] / "data" / "tables.json"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
T = {}


def rows(name, maxr=200, maxc=40):
    return [list(r) for r in wb[name].iter_rows(max_row=maxr, max_col=maxc, values_only=True)]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---- WPS 용접속도: [tmin, tmax, mm/s, pass] -------------------------------
def wps(name, c0, speed_off, pass_off):
    out = []
    for r in rows(name, 60, 32):
        mn, mx = num(r[c0 + 1]), num(r[c0 + 2])
        sp, ps = num(r[c0 + speed_off]), num(r[c0 + pass_off])
        if mn is None or mx is None or sp is None:
            continue
        if mx < mn:          # 마지막 행의 최대값 오타 보정
            mx = 99.0
        out.append([mn, mx, round(sp, 4), int(ps or 1)])
    return out


T["tackWeld"] = wps("Tack Welder WPS", 1, 6, 4)
T["insideWeld"] = wps("Inside Welder WPS", 0, 7, 5)
T["outsideWeld"] = wps("Outside Welder WPS", 1, 7, 5)

# ---- 1st-UT 두께별 절단속도 ----------------------------------------------
T["utCut"] = [
    [num(r[0]), num(r[1]), num(r[2])]
    for r in rows("1st-UT 두께별 절단 속도", 40, 6)
    if None not in (num(r[0]), num(r[1]), num(r[2]))
]

# ---- Edge Miller: 고속 Setting 값 / 메인 피딩기 전진거리 -------------------
T["emSpeed"] = [
    {"tmin": 0, "tmax": 18, "normal": 5, "high": 3},
    {"tmin": 18.001, "tmax": 100, "normal": 4, "high": 2},
]
T["emFeed"] = [[0, 12.9, 15250], [13, 16, 17250], [16.1, 19, 19250]]

# ---- Pre Bender 성형 피치 -------------------------------------------------
T["preBenderPitch"] = [[8, 25.4, 1450], [25.401, 99, 800]]

# ---- Press Bender X1 Side Press 횟수 (외경 inch) ---------------------------
px = {}
for r in rows("Press Bender(18M)", 60, 6):
    s = r[1]
    if isinstance(s, str) and re.match(r"^\d+", s.strip()):
        n = num(r[2])
        if n:
            px[int(re.match(r"^(\d+)", s.strip()).group(1))] = int(n)
T["pressX1"] = px

# ---- End-Facing 저속절삭 시간 [inch, tmin, tmax, sec] ----------------------
ef = []
for r in rows("End-Facing 저속절삭 시간", 200, 8):
    inch, trange, sec = num(r[0]), r[1], num(r[5])
    if inch is None or sec is None or not isinstance(trange, str):
        continue
    m = re.match(r"\s*(\d+(?:\.\d+)?)\s*~\s*(\d+(?:\.\d+)?)?", trange)
    if not m:
        continue
    ef.append([int(inch), float(m.group(1)), float(m.group(2)) if m.group(2) else 999.0, round(sec, 2)])
T["endFacing"] = ef
T["endFacingTC"] = {"클램프 교체(외경 변화)": 60, "우레탄 면판 교체": 30, "카트리지 교체": 10, "팁 교체": 5}

# ---- 수압: 외경별 충수시간 + 고정 상수 -------------------------------------
hf = {}
for r in wb["Hydraulic Tester 수압시간"].iter_rows(max_row=60, max_col=9, values_only=True):
    s, v = r[1], r[2]
    if isinstance(s, str) and re.match(r'^\d+"', s.strip()):
        try:
            hf[int(re.match(r"^(\d+)", s.strip()).group(1))] = float(v)
        except (TypeError, ValueError):
            pass
T["hydroFill"] = hf
T["hydroConst"] = {
    "pressureRise": 30, "deflate2nd": 180, "airVent": 120,
    "deflate2nd_36up": 300, "airVent_36up": 180, "faceplateChangeMin": 50,
}

# ---- 확관 1호기 다이 테이블: inch -> [[두께, step mm], ...] -----------------
dies, cur = {}, None
for r in rows("Expander(1호기)", 200, 10):
    inch = num(r[2])
    if inch:
        cur = int(inch)
    if cur is None:
        continue
    for c in r[3:9]:
        if not isinstance(c, str):
            continue
        m = re.search(r"(\d+(?:\.\d+)?)\s*\(?\s*\(?(\d+)\s*mm", c.replace("\n", " "))
        if m:
            dies.setdefault(cur, []).append([float(m.group(1)), int(m.group(2))])
T["expanderDie"] = {k: sorted(v) for k, v in dies.items()}
# 주: Expander(2호기) 시트는 원본이 비어 있어 1호기 다이 테이블을 공용으로 사용

# ---- 포장 마킹 -------------------------------------------------------------
T["packingMarking"] = {
    "markSpec": {"내면": 1, "외면": 1, "내/외면": 2},
    "markEnd": {"선단": 1, "후단": 1, "선/후단": 2},
}

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(json.dumps(T, ensure_ascii=False, indent=1), encoding="utf-8")
print(f"[ok] {OUT}")
for k, v in T.items():
    print(f"  {k:18s} {len(v) if hasattr(v, '__len__') else v}")
